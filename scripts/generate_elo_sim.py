import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_simulation():
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    calc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    result_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet 1: Single Match Simulator ---
    ws1 = wb.active
    ws1.title = "Single Match Simulator"

    # Column widths
    widths = {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 10, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 5, 'N': 10, 'O': 10, 'P': 10}
    for col, width in widths.items():
        ws1.column_dimensions[col].width = width

    # Header
    ws1.merge_cells('A1:P1')
    ws1['A1'] = "ELO Match Simulator (Grabbarnas Serie)"
    ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal="center")

    # Match Settings
    ws1['A3'] = "Match Settings"
    ws1['A3'].font = Font(bold=True)

    settings = [
        ("Score Type", "Sets", "Sets or Points"),
        ("Score Target (if Points)", 21, "Points needed to win (e.g. 21, 31)"),
        ("Is Tournament?", "No", "Tournament matches have 1.0 weight"),
        ("Is Singles?", "No", "1v1 matches count 0.5x weight"),
    ]

    for i, (label, val, desc) in enumerate(settings):
        row = 4 + i
        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = val
        ws1[f'C{row}'] = desc
        ws1[f'B{row}'].fill = input_fill
        ws1[f'B{row}'].border = border

    ws1['A8'] = "Team 1 Score"
    ws1['B8'] = 2
    ws1['B8'].fill = input_fill
    ws1['B8'].border = border

    ws1['A9'] = "Team 2 Score"
    ws1['B9'] = 0
    ws1['B9'].fill = input_fill
    ws1['B9'].border = border

    # Data Validations
    dv_score_type = DataValidation(type="list", formula1='"Sets,Points"', allow_blank=True)
    ws1.add_data_validation(dv_score_type)
    dv_score_type.add(ws1['B4'])

    dv_tournament = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws1.add_data_validation(dv_tournament)
    dv_tournament.add(ws1['B6'])

    dv_singles = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws1.add_data_validation(dv_singles)
    dv_singles.add(ws1['B7'])

    # Players Section
    ws1['A12'] = "Players Input"
    ws1['A12'].font = Font(bold=True)

    headers = ["Name", "Current ELO", "Games Played", "Team"]
    for i, h in enumerate(headers):
        cell = ws1.cell(row=13, column=i+1)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = border

    players = [
        ("Player 1A", 1000, 0, "Team 1"),
        ("Player 1B", 1000, 0, "Team 1"),
        ("Player 2A", 1000, 0, "Team 2"),
        ("Player 2B", 1000, 0, "Team 2"),
    ]

    for i, (name, elo, games, team) in enumerate(players):
        row = 14 + i
        ws1[f'A{row}'] = name
        ws1[f'B{row}'] = elo
        ws1[f'C{row}'] = games
        ws1[f'D{row}'] = team
        for col in 'ABC':
            ws1[f'{col}{row}'].fill = input_fill
            ws1[f'{col}{row}'].border = border
        ws1[f'D{row}'].border = border

    # Calculations Section
    calc_headers = ["Team Avg", "K-Factor", "Exp. Score", "Player Weight", "Eff. Weight", "M. Multiplier", "Match Weight"]
    for i, h in enumerate(calc_headers):
        cell = ws1.cell(row=13, column=i+6)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = border

    # Team averages
    ws1['F14'] = "=AVERAGE(B14,B15)"
    ws1['F15'] = "=F14"
    ws1['F16'] = "=AVERAGE(B16,B17)"
    ws1['F17'] = "=F16"

    # Match level calculations
    # User request: 2 set difference should be 1.1x. So > 2 sets = 1.2x.
    ws1['K14'] = '=1 + IF(ABS(B8-B9)>2, 0.2, IF(ABS(B8-B9)>0, 0.1, 0))'
    ws1['L14'] = '=IF(B7="Yes", 0.5, 1) * IF(B6="Yes", 1, IF(B4="Sets", IF(MAX(B8,B9)>=6, 1, 0.5), IF(B5>21, 1, 0.5)))'
    for i in range(15, 18):
        ws1[f'K{i}'] = '=K14'
        ws1[f'L{i}'] = '=L14'

    for i in range(4):
        row = 14 + i
        ws1[f'G{row}'] = f'=IF(C{row}<10, 40, IF(C{row}<30, 30, 20))'
        opp_avg = "F16" if i < 2 else "F14"
        ws1[f'H{row}'] = f'=1 / (1 + 10^(({opp_avg} - F{row}) / 300))'
        ws1[f'I{row}'] = f'=MIN(1.25, MAX(0.75, 1 + (F{row} - B{row}) / 800))'
        win_cond = "B8>B9" if i < 2 else "B9>B8"
        ws1[f'J{row}'] = f'=IF({win_cond}, I{row}, 1/I{row})'

        for col in 'FGHIJKL':
            ws1[f'{col}{row}'].fill = calc_fill
            ws1[f'{col}{row}'].border = border

    # Results
    res_headers = ["Result", "Delta", "New ELO"]
    for i, h in enumerate(res_headers):
        cell = ws1.cell(row=13, column=i+14)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = border

    for i in range(4):
        row = 14 + i
        win_cond = "B8>B9" if i < 2 else "B9>B8"
        ws1[f'N{row}'] = f'=IF({win_cond}, "Win", "Loss")'
        actual = f"IF({win_cond}, 1, 0)"
        ws1[f'O{row}'] = f'=ROUND(G{row} * K{row} * L{row} * J{row} * ({actual} - H{row}), 0)'
        ws1[f'P{row}'] = f'=B{row} + O{row}'

        ws1[f'N{row}'].border = border
        ws1[f'O{row}'].fill = result_fill
        ws1[f'O{row}'].border = border
        ws1[f'P{row}'].fill = result_fill
        ws1[f'P{row}'].border = border

    # --- Sheet 2: Scenarios ---
    ws2 = wb.create_sheet("Scenarios")
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15

    scenarios = [
        ("Scenario 1: Balanced", "Everyone starts at 1000 ELO. 2-0 Win."),
        ("Player", "Start ELO", "Games", "Opp. Avg", "Result", "Delta", "End ELO"),
        ("P1A", 1000, 0, 1000, "Win", 12, 1012),
        ("P1B", 1000, 0, 1000, "Win", 12, 1012),
        ("P2A", 1000, 0, 1000, "Loss", -12, 988),
        ("P2B", 1000, 0, 1000, "Loss", -12, 988),
        ("", "", "", "", "", "", ""),
        ("Scenario 2: Underdog", "Team 1 (Avg 800) vs Team 2 (Avg 1200). 2-1 Win for T1."),
        ("Player", "Start ELO", "Games", "Opp. Avg", "Result", "Delta", "End ELO"),
        ("P1A (800)", 800, 30, 1200, "Win", 19, 819),
        ("P1B (800)", 800, 30, 1200, "Win", 19, 819),
        ("P2A (1200)", 1200, 30, 800, "Loss", -19, 1181),
        ("P2B (1200)", 1200, 30, 800, "Loss", -19, 1181),
        ("", "", "", "", "", "", ""),
        ("Scenario 3: Carry", "T1: P1A(1400) & P1B(600) vs T2: Avg 1000. 2-0 Win for T1."),
        ("Player", "Start ELO", "Games", "Opp. Avg", "Result", "Delta", "End ELO"),
        ("P1A (1400)", 1400, 50, 1000, "Win", 4, 1404),
        ("P1B (600)", 600, 50, 1000, "Win", 11, 611),
        ("P2A (1000)", 1000, 50, 1000, "Loss", -7, 993),
        ("P2B (1000)", 1000, 50, 1000, "Loss", -7, 993),
    ]

    for r_idx, row_data in enumerate(scenarios, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.value = val
            if "Scenario" in str(val):
                cell.font = Font(bold=True, size=12)
            if val in ["Player", "Start ELO", "Games", "Opp. Avg", "Result", "Delta", "End ELO"]:
                cell.font = Font(bold=True)
                cell.fill = calc_fill
            cell.border = border

    # --- Sheet 3: Sequence Simulator ---
    ws3 = wb.create_sheet("Sequence Simulator")
    ws3['A1'] = "Sequence of Matches"
    ws3['A1'].font = Font(size=14, bold=True)

    ws3['A3'] = "Instructions: Enter match results row by row. Use the Single Match sheet to find deltas."
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 10

    seq_headers = ["Match #", "Player Name", "Start ELO", "Delta", "End ELO"]
    for i, h in enumerate(seq_headers):
        cell = ws3.cell(row=5, column=i+1)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = border

    for i in range(1, 11):
        row = 5 + i
        ws3[f'A{row}'] = f"Match {i}"
        ws3[f'B{row}'].fill = input_fill
        ws3[f'C{row}'] = f"=IF(B{row}=\"\", \"\", 1000)" if i == 1 else f"=E{row-1}"
        ws3[f'D{row}'].fill = input_fill
        ws3[f'E{row}'] = f"=IF(B{row}=\"\", \"\", C{row}+D{row})"
        for col in 'ABCDE':
            ws3[f'{col}{row}'].border = border

    # --- Sheet 4: MVP Simulator ---
    ws_mvp = wb.create_sheet("MVP Simulator")
    ws_mvp.column_dimensions['A'].width = 20
    ws_mvp.column_dimensions['B'].width = 15
    ws_mvp.column_dimensions['C'].width = 15
    ws_mvp.column_dimensions['D'].width = 15
    ws_mvp.column_dimensions['E'].width = 15
    ws_mvp.column_dimensions['F'].width = 15
    ws_mvp.column_dimensions['G'].width = 15
    ws_mvp.column_dimensions['H'].width = 20
    ws_mvp.column_dimensions['I'].width = 20

    ws_mvp.merge_cells('A1:I1')
    ws_mvp['A1'] = "MVP Scoring Simulator"
    ws_mvp['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_mvp['A1'].fill = header_fill
    ws_mvp['A1'].alignment = Alignment(horizontal="center")

    mvp_headers = ["Player Name", "ELO Gain", "Wins", "Games Played", "Win Rate", "Total ELO", "Eligibility", "Standard MVP Score", "Rolling MVP Score"]
    for i, h in enumerate(mvp_headers):
        cell = ws_mvp.cell(row=3, column=i+1)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = calc_fill

    for i in range(5):
        row = 4 + i
        ws_mvp[f'A{row}'] = f"Player {i+1}"
        ws_mvp[f'A{row}'].fill = input_fill
        ws_mvp[f'B{row}'] = 20
        ws_mvp[f'B{row}'].fill = input_fill
        ws_mvp[f'C{row}'] = 3
        ws_mvp[f'C{row}'].fill = input_fill
        ws_mvp[f'D{row}'] = 4
        ws_mvp[f'D{row}'].fill = input_fill
        ws_mvp[f'F{row}'] = 1100
        ws_mvp[f'F{row}'].fill = input_fill

        ws_mvp[f'E{row}'] = f'=IF(D{row}=0, 0, C{row}/D{row})'
        ws_mvp[f'G{row}'] = f'=IF(D{row}>=6, "Month OK", IF(D{row}>=3, "Evening OK", "Too few games"))'
        # Standard MVP Score = eloGain + (winRate * 15) + (games * 0.5)
        ws_mvp[f'H{row}'] = f'=B{row} + (E{row} * 15) + (D{row} * 0.5)'
        # Harmonized MVP Days Score is now the same as Standard MVP Score
        ws_mvp[f'I{row}'] = f'=H{row}'

        for col in 'ABCDEF':
            ws_mvp[f'{col}{row}'].border = border
        for col in 'GHI':
            ws_mvp[f'{col}{row}'].border = border
            ws_mvp[f'{col}{row}'].fill = result_fill

    # --- Sheet 5: MVP Scenarios ---
    ws_mvp_scen = wb.create_sheet("MVP Scenarios")
    ws_mvp_scen.column_dimensions['A'].width = 25
    ws_mvp_scen.column_dimensions['B'].width = 15
    ws_mvp_scen.column_dimensions['C'].width = 15
    ws_mvp_scen.column_dimensions['D'].width = 15
    ws_mvp_scen.column_dimensions['E'].width = 15
    ws_mvp_scen.column_dimensions['F'].width = 20
    ws_mvp_scen.column_dimensions['G'].width = 20

    mvp_scenarios = [
        ("Scenario 1: The Efficiency King", "High win rate, few games (Evening MVP focus)."),
        ("Player", "ELO Gain", "Wins", "Games", "Win Rate", "Std MVP Score", "Roll MVP Score"),
        ("Efficiency King", 40, 3, 3, 1.00, 56.50, 56.50),
        ("The Grinder", 30, 4, 6, 0.67, 43.05, 43.05),
        ("", "", "", "", "", "", ""),
        ("Scenario 2: The Workhorse", "High game count compensates for lower efficiency (Month MVP focus)."),
        ("Player", "ELO Gain", "Wins", "Games", "Win Rate", "Std MVP Score", "Roll MVP Score"),
        ("Workhorse", 50, 7, 10, 0.70, 65.50, 65.50),
        ("Elite Sprinter", 60, 5, 6, 0.83, 75.45, 75.45),
        ("", "", "", "", "", "", ""),
        ("Scenario 3: The Comeback", "Massive ELO gains from underdog wins."),
        ("Player", "ELO Gain", "Wins", "Games", "Win Rate", "Std MVP Score", "Roll MVP Score"),
        ("Underdog Hero", 80, 3, 4, 0.75, 93.25, 93.25),
        ("Consistent Pro", 40, 5, 5, 1.00, 57.50, 57.50),
    ]

    for r_idx, row_data in enumerate(mvp_scenarios, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_mvp_scen.cell(row=r_idx, column=c_idx)
            cell.value = val
            if "Scenario" in str(val):
                cell.font = Font(bold=True, size=12)
            if val in ["Player", "ELO Gain", "Wins", "Games", "Win Rate", "Std MVP Score", "Roll MVP Score"]:
                cell.font = Font(bold=True)
                cell.fill = calc_fill
            cell.border = border

    # --- Sheet 6: Documentation ---
    ws4 = wb.create_sheet("Documentation")
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 80

    doc = [
        ("Term", "Explanation"),
        ("Baseline ELO", "Starting rating for new players. Fixed at 1000."),
        ("K-Factor", "Determines volatility. <10 games: 40, 10-29 games: 30, 30+ games: 20."),
        ("Expected Score", "Win probability (0 to 1). Calculated using rating difference (divisor 300)."),
        ("Margin Multiplier", "Bonus for decisive wins. 1-2 sets = 1.1x multiplier. 3+ sets = 1.2x."),
        ("Match Length Weight", "Short games (sets <= 3 or points <= 21) = 0.5x. Tournament/Long = 1.0x."),
        ("Singles Match Weight", "1v1 matches count for 0.5x weight."),
        ("Player Weight", "Adjusts gain based on partner difference. Lower rated players get a boost (0.75x to 1.25x)."),
        ("Effective Weight", "Uses Player Weight for wins and 1/Weight for losses."),
        ("Guest Players", "The app ignores players marked as Guests. In this simulator, simply leave their ELO blank to exclude them from team averages."),
        ("", ""),
        ("MVP Scoring", ""),
        ("Standard MVP Score", "Used for Evening and Month MVP awards."),
        ("MVP Days Score", "Now harmonized with the Standard MVP Score."),
        ("Evening MVP Eligibility", "Must play at least 3 matches in the session."),
        ("Month MVP Eligibility", "Must play at least 6 matches in the rolling 30-day window."),
        ("", ""),
        ("Excel Formulas used:", ""),
        ("Expected Score", "=1 / (1 + 10^((Opponent_Avg - Own_Avg) / 300))"),
        ("Player Weight", "=MIN(1.25, MAX(0.75, 1 + (Team_Avg - Player_Elo) / 800))"),
        ("Margin Multiplier", "=1 + IF(ABS(Set_Diff)>2, 0.2, IF(ABS(Set_Diff)>0, 0.1, 0))"),
        ("Delta (ELO)", "=ROUND(K * MarginMult * MatchWeight * SinglesWeight * EffWeight * (Result - Expected), 0)"),
        ("Standard MVP Score", "=eloGain + (winRate * 15) + (gamesPlayed * 0.5)"),
        ("MVP Days Score", "=eloGain + (winRate * 15) + (gamesPlayed * 0.5)"),
    ]

    for r_idx, (term, expl) in enumerate(doc, 1):
        ws4[f'A{r_idx}'] = term
        ws4[f'B{r_idx}'] = expl
        ws4[f'A{r_idx}'].font = Font(bold=True)
        ws4[f'A{r_idx}'].border = border
        ws4[f'B{r_idx}'].border = border
        if term == "Term" or term == "MVP Scoring" or term == "Excel Formulas used:":
            ws4[f'A{r_idx}'].fill = header_fill
            ws4[f'A{r_idx}'].font = Font(color="FFFFFF", bold=True)
            ws4[f'B{r_idx}'].fill = header_fill
            ws4[f'B{r_idx}'].font = Font(color="FFFFFF", bold=True)

    wb.save("ELO_Simulation.xlsx")
    print("ELO_Simulation.xlsx created/updated successfully.")

if __name__ == "__main__":
    create_simulation()
